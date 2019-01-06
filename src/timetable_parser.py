from openpyxl import load_workbook
from openpyxl import styles
wb = load_workbook('Ist Year Sec A.xlsx', data_only = True)
sheet = wb['NIstYrSecA']

mergedCells = (sheet.merged_cells.ranges) #list with merged cell objects
dayCell_Bounds = dict()
c = 0
x = mergedCells[0] 
for cell in mergedCells:
	stringBounds = cell.ref #"A23:A30"
	splitColon = stringBounds.split(":")
	#extract alphabet and number from each element in splitColon
	alphabetA = splitColon[0][0] #assumption - alphabet is of length 1
	alphabetB = splitColon[1][0] #assumption - alphabet is of length 1
	if(alphabetA == alphabetB and alphabetA == 'A'):
		val = sheet[splitColon[0]].value #content of merged cell (should be always mon, tue, wed, thurs, fri and nothing else)
		dayCell_Bounds[val] = stringBounds

#calculate no. of columns(acc. to convention) in row 2 -> get no. of 30 minute time cells 
#assumption - cell[2][0] should be "day" cell

row_time = 2
max_col = sheet.max_column
time_cells = 0
for i in range(1, max_col + 1):
	cellObj = sheet.cell(row = row_time, column = i)
	if(cellObj.value is not None):
		time_cells += 1
		#print(cellObj.value)
time_cells -= 1
#time_cells is now the number of <30 minute time slots>
for k, v in dayCell_Bounds.items():
	mincol = 2
	maxcol = time_cells
	stringBounds = v
	splitColon = stringBounds.split(":")
	minrow = int(splitColon[0][1:]) 
	maxrow = int(splitColon[1][1:])
	curRow = minrow
	curCol = mincol
	print(minrow, maxrow)
	print(mincol, maxcol)
	while(curRow <= maxrow):
		curCol = mincol
		while(curCol <= maxcol):
			cellObj = sheet.cell(row = curRow, column = curCol)
			if(cellObj.value is not None):
				isMerged = 0
				duration = 1
				for i in mergedCells:
					candidateMergeCell = i
					"""check if current cell is part of candidate merge cell"""
					condition1 = (candidateMergeCell.min_row <= curRow) and (candidateMergeCell.max_row >= curRow)
					condition2 = (candidateMergeCell.min_col <= curCol) and (candidateMergeCell.max_col >= curCol)
					if(condition1 and condition2):
						duration = candidateMergeCell.max_col - candidateMergeCell.min_col + 1 
						isMerged = 1
						break
				parseString = cellObj.value
				originalColor = (cellObj.fill.start_color.index)
				#to find time range of a course, follow same color path
				tempCol = curCol + 1
				#if not part of merged cell, use size of same color sequence to find time duration
				if(not isMerged):
					while(tempCol <= maxcol):
						cellObj2 = sheet.cell(row = curRow, column = tempCol)
						if(cellObj2.fill.start_color.index == originalColor):
							duration += 1
							tempCol += 1
						else:
							break
				durationArr = [sheet.cell(row = 2, column = i).value for i in range(curCol, curCol + duration)]
				starttime = durationArr[0].split("-")[0]
				endtime = durationArr[-1].split("-")[-1]
				print(parseString, starttime + "-" + endtime)
			curCol += 1
		curRow += 1
startrow = None
endrow = sheet.max_row
startcol = 1
endcol = sheet.max_column
for cell in mergedCells:
	stringBounds = cell.ref #"A23:A30"
	splitColon = stringBounds.split(":")
	#extract alphabet and number from each element in splitColon
	string = sheet[splitColon[0]].value
	if(string is None):
		continue
	if(string.lower().find("duration") != -1):
		startrow = int(splitColon[0][1:]) + 1
		break
print(startrow)
file = open("courseList.txt", "w")
for i in range(startrow, endrow + 1):
	for j in range(startcol, endcol + 1):
		cell = sheet.cell(row = i, column = j)
		if(cell.value is not None):
			print(cell.value, type(cell.value))
			file.write(str(cell.value) + ";")
	file.write("\n")
file.close()