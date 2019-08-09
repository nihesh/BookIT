from openpyxl import load_workbook
from openpyxl import styles

try:

	CSV_DATA = set() 	# final data in tuple form that has to be wriiten to csv in that order
	booking = []		# Every row is a 3 tuple, consisting of parseString, start time and end time - per sheet data
	course_description = []		# Per-sheet data

	def process_cell_content(s):

		"""
		Processes the raw string s, which is the content of a cell
		"""
		data = {}

		s = s.replace("~"," ~ ")
		s = s.replace("  "," ")

		# Remove content written within brackets
		while(s.find("(") != -1):
			s = s[:s.find("(")]+s[s.find(")")+1:]

		# Invalid Cell
		if(s.find("~") == -1):
			return -1

		# Separate out room information from string and remove spaces
		rooms = s[s.find("~")+1:]
		rooms = rooms.replace(" ","")
		rooms = rooms.split(",")
		s = s[:s.find("~")]

		s = s.split(" ")
		new_list = []
		for val in s:
			if(s!=""):
				new_list.append(val)

		# Check if the cell corresponds to lab, tut or lecture
		if(new_list[0].lower() == "lab"):
			data["type"] = "Lab"
			new_list.pop(0)
		elif(new_list[0].lower() == "tut"):
			data["type"] = "Tutorial"
			new_list.pop(0)
		else:
			data["type"] = "Lecture"

		data["rooms"] = rooms
		data["group"] = "0"
		data["course_abbr"] = new_list[0].upper()
		i = 0
		while(i<len(new_list) and new_list[i].lower()!="gp"):
			i+=1
		if(i!=len(new_list)):
			data["group"] = new_list[i+1].replace("/"," ").replace(","," ")
			data["group"] = "0"
			new_list.pop(i)
			new_list.pop(i)

		# get group info and lecture name

		return data

	def process_time(t):

		t = t.replace(".",":")
		if(t.find(":") == -1):
			t+=":00"
		while(t.find(":")<2):
			t = "0"+t
		hours = int(t[:t.find(":")])
		if(hours >= 7 and hours <= 11):	
			t+="AM"
		else:
			t+="PM"
		t = t.replace(":","")

		return t

	def process_raw_bookings():

		"""
		Reads booking, converts raw data to a structured format and overwrites the content
		"""

		global booking, course_description, CSV_DATA

		processed_booking = []

		for entry in booking:
			processed_booking.append([process_cell_content(entry[0]), process_time(entry[1]), process_time(entry[2]), process_day(entry[3])])
		booking = processed_booking

		course_description.pop(0)
		data = {}
		
		for entry in course_description:
			data[entry[5]] = {}
			data[entry[5]]["core/elective"] = entry[0]
			data[entry[5]]["course_code"] = entry[1]
			data[entry[5]]["course_name"] = entry[2]
			data[entry[5]]["instructor"] = entry[3].replace(",","+")
			data[entry[5]]["credit"] = entry[4]

		for entry in booking:

			# skip invalid cell
			if(entry[0] == -1):
				continue

			for room in entry[0]["rooms"]:
				try:
					room = room.strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					course_abbr = entry[0]["course_abbr"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					message = entry[0]["type"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					group = entry[0]["group"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					start_time = entry[1].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					end_time = entry[2].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					day = entry[3].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					course_code = data[course_abbr]["course_code"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					mandatory_elective = data[course_abbr]["core/elective"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					instructor = data[course_abbr]["instructor"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					course_name = (data[course_abbr]["course_name"]+"-"+instructor).strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					credits = data[course_abbr]["credit"].strip(" ").strip("\n").rstrip(" ").rstrip("\n")
					CSV_DATA.add((mandatory_elective, course_code, course_name, instructor, credits, course_abbr, day, start_time, end_time, group, message, room))
				except KeyError:
					print("Course code",course_abbr,"not found")
					raise Exception

		return

	def process_day(day):

		# Remove blank spaces
		day = day.replace(" ","")
		Data = {"Mon":"Monday", "Tue":"Tuesday","Tues":"Tuesday","Wed":"Wednesday","Thurs":"Thursday","Thu":"Thursday","Thur":"Thursday","Fri":"Friday","Sat":"Saturday","Sun":"Sunday"}
		return Data[day]

	def load_raw_data():

		"""
		Reads the raw databases and populates booking list
		"""

		global booking, course_description


		wb = load_workbook('Timetable.xlsx', data_only = True)
		for sheet in wb.sheetnames:
			sheet = wb[sheet]
			print("Processing sheet", sheet)
			booking = []
			course_description = []

			mergedCells = (sheet.merged_cells.ranges) #list with merged cell objects
			dayCell_Bounds = dict()
			c = 0
			x = mergedCells[0] 
			for cell in mergedCells:
				stringBounds = cell.ref #"A23:A30"
				splitColon = stringBounds.split(":")
				#extract alphabet and number from each element in splitColon
				alphabetA = splitColon[0][0] #assumption - alphabet is of length 1
				alphabetB = splitColon[1][0] #assumption - alphabet is of length 1
				if(alphabetA == alphabetB and alphabetA == 'A'):
					val = sheet[splitColon[0]].value #content of merged cell (should be always mon, tue, wed, thurs, fri and nothing else)
					dayCell_Bounds[val] = stringBounds

			#calculate no. of columns(acc. to convention) in row 2 -> get no. of 30 minute time cells 
			#assumption - cell[2][0] should be "day" cell

			row_time = 2
			max_col = sheet.max_column
			time_cells = 0
			for i in range(1, max_col + 1):
				cellObj = sheet.cell(row = row_time, column = i)
				if(cellObj.value is not None):
					time_cells += 1
					#print(cellObj.value)
			time_cells -= 1
			#time_cells is now the number of <30 minute time slots>
			for k, v in dayCell_Bounds.items():
				mincol = 2
				maxcol = time_cells+1
				stringBounds = v
				splitColon = stringBounds.split(":")
				minrow = int(splitColon[0][1:]) 
				maxrow = int(splitColon[1][1:])
				curRow = minrow
				curCol = mincol
				# print(minrow, maxrow)
				# print(mincol, maxcol)
				while(curRow <= maxrow):
					curCol = mincol
					while(curCol <= maxcol):
						cellObj = sheet.cell(row = curRow, column = curCol)
						if(cellObj.value is not None):
							isMerged = 0
							duration = 1
							for i in mergedCells:
								candidateMergeCell = i
								"""check if current cell is part of candidate merge cell"""
								condition1 = (candidateMergeCell.min_row <= curRow) and (candidateMergeCell.max_row >= curRow)
								condition2 = (candidateMergeCell.min_col <= curCol) and (candidateMergeCell.max_col >= curCol)
								if(condition1 and condition2):
									duration = candidateMergeCell.max_col - candidateMergeCell.min_col + 1 
									isMerged = 1
									break
							parseString = cellObj.value.replace("\n", "")
							originalColor = (cellObj.fill.start_color.index)
							#to find time range of a course, follow same color path
							tempCol = curCol + 1
							#if not part of merged cell, use size of same color sequence to find time duration
							if(not isMerged):
								while(tempCol <= maxcol):
									cellObj2 = sheet.cell(row = curRow, column = tempCol)
									if(cellObj2.fill.start_color.index == originalColor and cellObj2.value is None):
										duration += 1
										tempCol += 1
									else:
										break
							durationArr = [sheet.cell(row = 2, column = i).value.replace("\n", "") for i in range(curCol, curCol + duration)]
							starttime = durationArr[0].split("-")[0]
							endtime = durationArr[-1].split("-")[-1]
							# Modified by Nihesh - populating data into global variable
							booking.append([parseString,starttime,endtime,k])

						curCol += 1
					curRow += 1
			startrow = None
			endrow = sheet.max_row
			startcol = 1
			endcol = sheet.max_column
			for cell in mergedCells:
				stringBounds = cell.ref #"A23:A30"
				splitColon = stringBounds.split(":")
				#extract alphabet and number from each element in splitColon
				string = sheet[splitColon[0]].value
				if(string is None):
					continue
				if(string.lower().find("duration") != -1):
					startrow = int(splitColon[0][1:]) + 1
					break

			for i in range(startrow, endrow + 1):
				now = []
				for j in range(startcol, endcol + 1):
					cell = sheet.cell(row = i, column = j)
					if((cell.value is not None) and (str(cell.value).strip() != "")):
						now.append(str(cell.value).strip(" ").rstrip(" ").replace("\n", ""))
				if(len(now) >= 6):
					now[5] = now[5].upper()

				# Modified by Nihesh - populating data into global variable
				if(len(now)!=0):
					if(len(now) == 6):
						now[0] = "Elective"
						now[2] = now[2].replace(",", " ")
					else:
						now[1] = now[1].replace(",", " ")
						now = ["Elective"] + now	
					course_description.append(now)

			process_raw_bookings()

		return

	def generate_csv():

		global CSV_DATA

		file = open("TimeTable.csv","w")
		file.write("Mandatory/Elective,Course Code,Course Name,Instructor,Credits,Acronym,Day,Start Time,End Time,Group,Message,Venue\n")
		for entry in CSV_DATA:
			for i in range(len(entry)):
				cell = entry[i]
				if(i!=len(entry)-1):
					file.write(cell+",")
				else:
					file.write(cell+"\n")

		file.close()

	if(__name__ == "__main__"):

		# CSV header
		
		print("Loading Raw data from excel sheet")
		load_raw_data()
		print("Loaded Data")

		print("Writing to csv")
		generate_csv()
		print("Generation complete!")

except Exception as e:

	print("Error occurred. Exiting")

	print("\nDetailed Error")
	print(str(e))

	import traceback
	traceback.print_exc()