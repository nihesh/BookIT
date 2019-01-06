from openpyxl import load_workbook
from openpyxl import styles

booking = []		# Every row is a 3 tuple, consisting of parseString, start time and end time

def process_cell_content(s):

	"""
	Processes the raw string s, which is the content of a cell
	"""

	data = []

	# Remove content written within brackets
	while(s.find("(") != -1):
		s = s[:s.find("(")]+s[s.find(")")+1:]

	s = s.split(" ")
	new_list = []
	for val in s:
		if(s!=""):
			new_list.append(val)

	if(new_list[0].lower() == "lab" or new_list[0].lower() == "tut"):
		data.append(new_list[0].title())
	else:
		data.append("Lecture")

	return s

def process_time(t):

	t = t.replace(".",":")
	if(t.find(":") == -1):
		t+=":00"
	while(t.find(":")<2):
		t = "0"+t
	hours = int(t[:t.find(":")])
	if(hours >= 7 and hours <= 11):
		t+="AM"
	else:
		t+="PM"

	return t

def process_raw_bookings():

	"""
	Reads booking, converts raw data to a structured format and overwrites the content
	"""

	global booking

	processed_booking = []

	for entry in booking:
		processed_booking.append([process_cell_content(entry[0]), process_time(entry[1]), process_time(entry[2])])
	booking = processed_booking

	return

def load_raw_data():

	"""
	Reads the raw databases and populates booking list
	"""

	global booking


	wb = load_workbook('Ist Year Sec A.xlsx', data_only = True)
	sheet = wb['NIstYrSecA']

	mergedCells = (sheet.merged_cells.ranges) #list with merged cell objects
	dayCell_Bounds = dict()
	c = 0
	x = mergedCells[0] 
	for cell in mergedCells:
		stringBounds = cell.ref #"A23:A30"
		splitColon = stringBounds.split(":")
		#extract alphabet and number from each element in splitColon
		alphabetA = splitColon[0][0] #assumption - alphabet is of length 1
		alphabetB = splitColon[1][0] #assumption - alphabet is of length 1
		if(alphabetA == alphabetB and alphabetA == 'A'):
			val = sheet[splitColon[0]].value #content of merged cell (should be always mon, tue, wed, thurs, fri and nothing else)
			dayCell_Bounds[val] = stringBounds

	#calculate no. of columns(acc. to convention) in row 2 -> get no. of 30 minute time cells 
	#assumption - cell[2][0] should be "day" cell

	row_time = 2
	max_col = sheet.max_column
	time_cells = 0
	for i in range(1, max_col + 1):
		cellObj = sheet.cell(row = row_time, column = i)
		if(cellObj.value is not None):
			time_cells += 1
			#print(cellObj.value)
	time_cells -= 1
	#time_cells is now the number of <30 minute time slots>
	for k, v in dayCell_Bounds.items():
		mincol = 2
		maxcol = time_cells
		stringBounds = v
		splitColon = stringBounds.split(":")
		minrow = int(splitColon[0][1:]) 
		maxrow = int(splitColon[1][1:])
		curRow = minrow
		curCol = mincol
		# print(minrow, maxrow)
		# print(mincol, maxcol)
		while(curRow <= maxrow):
			curCol = mincol
			while(curCol <= maxcol):
				cellObj = sheet.cell(row = curRow, column = curCol)
				if(cellObj.value is not None):
					isMerged = 0
					duration = 1
					for i in mergedCells:
						candidateMergeCell = i
						"""check if current cell is part of candidate merge cell"""
						condition1 = (candidateMergeCell.min_row <= curRow) and (candidateMergeCell.max_row >= curRow)
						condition2 = (candidateMergeCell.min_col <= curCol) and (candidateMergeCell.max_col >= curCol)
						if(condition1 and condition2):
							duration = candidateMergeCell.max_col - candidateMergeCell.min_col + 1 
							isMerged = 1
							break
					parseString = cellObj.value
					originalColor = (cellObj.fill.start_color.index)
					#to find time range of a course, follow same color path
					tempCol = curCol + 1
					#if not part of merged cell, use size of same color sequence to find time duration
					if(not isMerged):
						while(tempCol <= maxcol):
							cellObj2 = sheet.cell(row = curRow, column = tempCol)
							if(cellObj2.fill.start_color.index == originalColor):
								duration += 1
								tempCol += 1
							else:
								break
					durationArr = [sheet.cell(row = 2, column = i).value for i in range(curCol, curCol + duration)]
					starttime = durationArr[0].split("-")[0]
					endtime = durationArr[-1].split("-")[-1]
					booking.append([parseString,starttime,endtime])
					print(process_cell_content(parseString), process_time(starttime) + "-" + process_time(endtime))
				curCol += 1
			curRow += 1
	startrow = None
	endrow = sheet.max_row
	startcol = 1
	endcol = sheet.max_column
	for cell in mergedCells:
		stringBounds = cell.ref #"A23:A30"
		splitColon = stringBounds.split(":")
		#extract alphabet and number from each element in splitColon
		string = sheet[splitColon[0]].value
		if(string is None):
			continue
		if(string.lower().find("duration") != -1):
			startrow = int(splitColon[0][1:]) + 1
			break
	print(startrow)
	file = open("courseList.txt", "w")
	for i in range(startrow, endrow + 1):
		for j in range(startcol, endcol + 1):
			cell = sheet.cell(row = i, column = j)
			if(cell.value is not None):
				print(cell.value, type(cell.value))
				file.write(str(cell.value) + ";")
		file.write("\n")
	file.close()

if(__name__ == "__main__"):


	print("Loading Raw data from excel sheet\n")
	load_raw_data()
	print("Loaded Data\n")

	print("Postprocessing bookings\n")
	# process_raw_bookings()
	print("Booking Postprocessing complete")
